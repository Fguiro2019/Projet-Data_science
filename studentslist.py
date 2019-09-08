# Fichier créé par Fadel Guiro le 08/09/2019
# Dernière modification  08/09/2019 à 20:25


# On importe la library openPyXL
import openpyxl


def get_sheet(filename, nb_sheet):
    """
    Cette fonction ouvre un fichier excel(Eleves.xlsx) et retourne une feuille(la premiere) du classeur
    On pourrait ameliorer pourqu'elle retourne une feuille donnee d'un classeur donne
    :param filename: Nom du Fichier à ouvrir
    :param nb_sheet: Numéro de la feuille  à retourner. Ce numéro commence à portir de 0, qui est la
    premiere feuille du classeur
    :return: Sheet la feuille
    """
    # Ouvrir le fichier excel
    excel_file = openpyxl.load_workbook(filename)
    try:
        # Acceder à une feuille
        return excel_file.worksheets[nb_sheet]
    except IndexError:
        raise ("Le numéro de la feuille n'existe pas. Veuillez compter à partir de 0 "
               "la feuille que vous souhaitez ouvrir.")


def average_list(sheet):
    """
    Elle permet d'obtenir la liste des eleves qui ont eu la moyenne
    :param sheet: Sheet, La feuille qui contient les donnees
    :return: Une liste D2 contenant les eleves qui ont eu la moyenne
    """
    result = []
    # On recuperer toutes les lignes de la feuille sous forme de liste de tuple
    rows = list(sheet.rows)
    # On recupere la valeur de la premiere ligne, On l'insere dans result
    result.append([rows[0][index].value for index in range(len(rows[0]))])
    del rows[0]
    for row in rows:
        row = [row[index].value for index in range(len(row))]
        average, tmp = str(row[3]).split('/')
        average = int(average)
        if average > 9:
            result.append(row)
    return result


def old_list(sheet):
    """
    Fonction permettant de creer une liste contenant les eleves qui ont plus de 20 ans
    :param sheet: Sheet, La feuille qui contient les donnees
    :return: Une liste D2 contenant les eleves qui ont plus de 20 ans
    """
    result = []
    rows = list(sheet.rows)
    result.append([rows[0][index].value for index in range(len(rows[0]))])
    del rows[0]
    for row in rows:
        row = [row[index].value for index in range(len(row))]
        if int(row[4]) > 20:
            result.append(row)
    return result


def get_school_average(sheet):
    """
    Fonction retournant la moyenne de l'ecole
    :param sheet: Sheet, La feuille qui contient les donnees
    :return: un reel contenant la moyenne de l'ecole
    """
    result = 0
    rows = list(sheet.rows)
    del rows[0]
    for row in rows:
        average, tmp = str(row[3].value).split('/')
        result = result + int(average)
    result = result / len(rows)
    result, tmp = str(result).split('.')
    result = result + '.' + tmp[:2] + '/20'
    return result


def get_percentage_girls_and_boys(sheet):
    """
    Fonction retournant le pourcentage de filles et de garcon
    :param sheet: Sheet, La feuille qui contient les donnees
    :return: une liste contenant le pourcentage de fille et de garcon
    """
    nb_girls, nb_boys = 0, 0
    rows = list(sheet.rows)
    del rows[0]
    for row in rows:
        sex = row[7].value
        if sex == 'F':
            nb_girls = nb_girls + 1
        else:
            nb_boys = nb_boys + 1
    nb_girls = (nb_girls / len(rows))*100
    nb_girls, tmp = str(nb_girls).split('.')
    nb_girls = nb_girls + '.' + tmp[:2]
    nb_girls = float(nb_girls)
    nb_boys = 100 - nb_girls
    nb_girls = "".join([str(nb_girls), '%'])
    nb_boys = "".join([str(nb_boys), '%'])
    return [nb_girls, nb_boys]


def get_best_region(sheet):
    """
    Fonction retournant la meilleurs region (region enregistrant la plus forte moyenne)
    :param sheet: Sheet, La feuille qui contient les donnees
    :return: Un string contenant la meilleurs region
    """

    rows = list(sheet.rows)
    del rows[0]
    region_result = rows[0][5].value
    average_result = rows[0][3].value
    average_result, tmp = str(average_result).split('/')
    average_result = int(average_result)
    for row in rows:
        average, tmp = str(row[3].value).split('/')
        average = int(average)
        if average_result < average:
            average_result = average
            region_result = row[5].value
    return region_result


def create_file(matrice_of_data, file_out):
    """
    C'est une fonction qui cree un fichier excel contenant le contenu de matrice_of_data.
    :param matrice_of_data: Une liste contenant les donnees a inserer dans le fichier
    :param file_out: Nom du fichier de sortie
    :return:
    """
    letters = [chr(index) for index in range(65, 65 + len(matrice_of_data[0]))]
    # Ici nous allons creer des classeur
    wb = openpyxl.Workbook()

    sheet = wb.active
    sheet.title = 'data'
    for i in range(len(matrice_of_data)):
        for letter in letters:
            sheet[letter + str(i + 1)] = matrice_of_data[i][ord(letter) - 65]
    wb.save(file_out)
    print('Creation du fichier ' + file_out + ' réussie!')


def create_statics_file(sheet, file_out):
    """
    Fonction permettant de creer un fichier excel contenant
    :param sheet: Une liste contenant les donnees a inserer dans le fichier
    :param file_out: Nom du fichier de sortie contenant les statistique
    :return: None
    """
    statistics = [get_school_average(sheet)]
    statistics.extend(get_percentage_girls_and_boys(sheet))
    statistics.append(get_best_region(sheet))

    wb = openpyxl.Workbook()

    sheet = wb.active
    sheet.title = 'data'
    sheet['A1'] = "Moyenne de l'ecole"
    sheet['B1'] = "Pourcentage de Fille"
    sheet['C1'] = "Pourcentage de garcon"
    sheet['D1'] = "Meilleurs region"

    sheet['A2'] = statistics[0]
    sheet['B2'] = statistics[1]
    sheet['C2'] = statistics[2]
    sheet['D2'] = statistics[3]

    wb.save(file_out)
    print('Creation du fichier ' + file_out + ' réussie!')



Sheet = get_sheet('Eleves.xlsx', 0)
Result = average_list(Sheet)
create_file(Result, 'Moyenne.xlsx')
Result = old_list(Sheet)
create_file(Result, 'Eleves_plus_de_20_ans.xlsx')
create_statics_file(Sheet, 'Statiques.xlsx')
